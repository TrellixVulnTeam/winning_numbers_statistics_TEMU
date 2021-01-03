import openpyxl
import random


def countFrequency(a_list):
    # Creating an empty dictionary
    freq = {}
    for items in a_list:
        freq[items] = a_list.count(items)

    # Sorted the order of frequency
    return sorted(freq.items(), key=lambda x: x[1], reverse=True)


def random_number_line(a_list):
    num_count = 0
    result = []
    while num_count != 5:
        val = random.choice(a_list)

        if not result:
            result.append(val)
        else:
            while val in result:
                val = random.choice(a_list)
            result.append(val)
        num_count = num_count + 1

    result.sort()
    result.append(random.randint(1, 26))
    return result


def export_file(file_name, myList):
    sorted_order = countFrequency(myList)

    # Call a Workbook() function of openpyxl
    # to create a new blank Workbook object
    wb = openpyxl.Workbook()

    # Get workbook active sheet
    # from the active attribute
    sheet = wb.active

    increment = 1
    for counter in sorted_order:
        output = "{} {}"
        temp = output.format(*counter).split()

        # using sheet object's cell() method.
        c1 = sheet.cell(row=increment, column=1)

        # writing values to cells
        c1.value = int(temp[0])

        c2 = sheet.cell(row=increment, column=2)
        c2.value = int(temp[1])

        increment = increment + 1

        # Anytime you modify the Workbook object
        # or its sheets and cells, the spreadsheet
        # file will not be saved until you call
        # the save() workbook method.
        wb.save(file_name)


def import_file(file_name):
    myList = []

    # workbook object is created
    wb_obj = openpyxl.load_workbook(file_name)

    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row

    num_data = 0
    num_line = 0

    # Loop will print all values
    # of first column
    for i in range(1, m_row + 1):
        # Check num of lines
        num_line = num_line + 1

        # Read line by line
        cell_obj = sheet_obj.cell(row=i, column=1)

        # Remove all whitespace in string
        format_num = cell_obj.value.replace(" ", "")

        count = 0
        word = ""

        for character in format_num:
            # Concatenate string
            word = word + character
            count = count + 1

            # Combine to a number
            if count == 2:
                count = 0
                myList.append(int(word))
                num_data = num_data + 1
                word = ""

    print("Number of line parsed:", num_line)
    print("Number of imported numbers:", num_data)
    return myList


def import_edited_file(file_name):
    myList = []

    # workbook object is created
    wb_obj = openpyxl.load_workbook(file_name)

    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row

    num_line = 0

    # Loop will print all values
    # of first column
    for i in range(1, int((m_row + 1) / 2)):
        # Check num of lines
        num_line = num_line + 1

        # Read line by line
        cell_obj = sheet_obj.cell(row=i, column=1)

        # Remove all whitespace in string
        val = int(cell_obj.value)

        myList.append(val)

    print("Number of line parsed:", num_line)
    return myList


# Driver function
if __name__ == "__main__":
    # Give the location of the file
    my_list = import_edited_file("output/statistic_output.xlsx")
    y = 0
    while y != 10:
        print("Random line: ", y, random_number_line(my_list))
        y = y + 1
